from django import forms
from django.contrib import admin, messages
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.utils import timezone
from django.utils.timezone import localtime
from django.urls import path
from django.http import HttpResponseRedirect
from django.shortcuts import render
import csv
import io
import logging
import openpyxl
from .models import EmailTemplate, Campaign, CampaignTarget, LessonProgress, QuizAttempt

logger = logging.getLogger(__name__)

# ── Custom admin form — routes smtp_password through the model property ────────

class CampaignAdminForm(forms.ModelForm):
    """
    Replaces the raw _smtp_password field with a proper password input
    that encrypts via the model's smtp_password property setter.
    Leave blank to keep the existing password.
    """
    smtp_password = forms.CharField(
        widget=forms.PasswordInput(render_value=False),
        required=False,
        help_text='Leave blank to keep the existing password.',
        label='SMTP Password',
    )
    csv_file = forms.FileField(
        required=False,
        label='Upload Targets (CSV)',
        help_text=(
            'Optional. Upload a CSV file to add targets immediately '
            'after saving. '
            "Required column: 'email' (case-insensitive). "
            "Optional columns: 'full_name', 'department', 'position', 'business_unit', 'manager', 'manager_email'."
        ),
    )

    class Meta:
        model  = Campaign
        fields = '__all__'
        exclude = ['_smtp_password']

    def save(self, commit=True):
        instance = super().save(commit=False)
        new_password = self.cleaned_data.get('smtp_password', '')
        if new_password:
            instance.smtp_password = new_password
        elif not instance.pk:
            instance.smtp_password = ''
        # NOTE: CSV processing is handled in CampaignAdmin.save_model()
        # because Django admin calls form.save(commit=False), so the
        # commit=True block here never runs during a normal admin save.
        if commit:
            instance.save()
            self.save_m2m()
        return instance

    def _process_file(self, campaign, upload_file):
        """
        Parse uploaded CSV or XLSX and create CampaignTarget rows.
        Returns (created, skipped, errors) counts.
        """
        filename = upload_file.name.lower()
        rows = []

        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            rows = self._read_xlsx(upload_file)
        else:
            rows = self._read_csv(upload_file)

        created = skipped = errors = 0
        for row in rows:
            email = row.get('email', '').strip().lower()
            if not email or '@' not in email:
                errors += 1
                continue
            _, was_created = CampaignTarget.objects.get_or_create(
                campaign=campaign,
                email=email,
                defaults={
                    'full_name': row.get('full_name', '').strip(),
                    'department': row.get('department', '').strip(),
                    'position': row.get('position', '').strip(),
                    'business_unit': row.get('business_unit', '').strip(),
                    'manager': row.get('manager', '').strip(),
                    'manager_email': row.get('manager_email', '').strip(),
                },
            )
            if was_created:
                created += 1
            else:
                skipped += 1

        logger.info(
            f'File upload for campaign "{campaign.name}": '
            f'{created} created, {skipped} skipped, {errors} invalid.'
        )
        self._csv_result = (created, skipped, errors)

    def _read_csv(self, upload_file):
        """Read CSV file and return list of normalised dicts."""
        # Read all bytes first, then decode — avoids the empty-read bug
        raw = upload_file.read()
        try:
            text = raw.decode('utf-8-sig')   # utf-8-sig strips BOM if present
        except UnicodeDecodeError:
            text = raw.decode('latin-1')

        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            # Normalise keys to lowercase stripped strings
            rows.append({
                k.strip().lower(): (v or '') for k, v in row.items()
            })
        return rows

    def _read_xlsx(self, upload_file):
        """Read XLSX file and return list of normalised dicts."""
        wb = openpyxl.load_workbook(upload_file, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # First row is the header
        try:
            headers = [str(h).strip().lower() if h else '' for h in next(rows_iter)]
        except StopIteration:
            return []

        rows = []
        for row in rows_iter:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val).strip() if val is not None else ''
            rows.append(row_dict)

        wb.close()
        return rows


# ── Email Template Admin ───────────────────────────────────────────────────────

class EmailTemplateAdminForm(forms.ModelForm):
    body_html = forms.CharField(
        widget=forms.Textarea(attrs={
            'rows': 30,
            'cols': 80,
            'style': 'font-family:monospace;width:100%;',
        }),
        label='Email Body (HTML)',
        help_text=mark_safe(
            '<strong>Plain text mode:</strong> Just type normally. '
            'Press Enter for a new line, press Enter twice for a new paragraph.<br><br>'
            '<strong>Clickable link with custom text:</strong> '
            'Use <code>[display text](url)</code> syntax — e.g. '
            '<code>[click here]({{ phishing_link }})</code><br><br>'
            '<strong>HTML mode:</strong> Write full HTML for custom layouts. '
            'Auto-detected when any HTML tags are present.<br><br>'
            '<strong>Available variables:</strong><br>'
            '<code>{{ target_name }}</code> &nbsp; '
            '<code>{{ target_email }}</code> &nbsp; '
            '<code>{{ target_department }}</code> &nbsp; '
            '<code>{{ phishing_link }}</code> &nbsp; '
            '<code>{{ company_name }}</code> &nbsp; '
            '<code>{{ campaign_name }}</code><br><br>'
            '<strong>Signature image:</strong> Upload an image below — it will be '
            'automatically appended to the bottom of every email sent with this template.'
        ),
    )

    class Meta:
        model  = EmailTemplate
        fields = '__all__'


@admin.register(EmailTemplate)
class EmailTemplateAdmin(admin.ModelAdmin):
    form = EmailTemplateAdminForm
    list_display  = ('name', 'sender_name', 'subject', 'company_name', 'created_by', 'created_at')
    list_filter   = ('created_at',)
    search_fields = ('name', 'subject', 'company_name', 'sender_name')
    readonly_fields = ('created_at', 'updated_at')

    fieldsets = (
        ('Template Info', {
            'fields': ('name', 'sender_name', 'subject', 'company_name', 'created_by')
        }),
        ('Email Body', {
            'fields': ('body_html',),
            'description': (
                'Available variables: {{ target_name }}, {{ target_email }}, '
                '{{ target_department }}, {{ phishing_link }}, '
                '{{ company_name }}, {{ campaign_name }}'
            ),
        }),
        ('Signature Image', {
            'fields': ('signature_image',),
            'description': (
                'Optional. Upload a signature image (e.g. company logo, '
                'sender signature graphic). It will be embedded inline at the '
                'bottom of every email sent using this template. '
                'Supported formats: PNG, JPG, GIF. Recommended max width: 400px.'
            ),
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',),
        }),
    )


# ── Campaign Target Inline ─────────────────────────────────────────────────────

class CampaignTargetInline(admin.TabularInline):
    model   = CampaignTarget
    extra   = 1
    fields  = (
        'email', 'full_name', 'department', 'position', 'business_unit',
        'manager_email', 'email_sent_at', 'email_failed', 'link_clicked_at',
        'lms_completed_at', 'quiz_score',
    )
    readonly_fields = (
        'token', 'email_sent_at', 'email_failed', 'email_error',
        'link_clicked_at', 'click_ip', 'lms_started_at',
        'lms_completed_at', 'quiz_score',
    )
    show_change_link = True


# ── Campaign Admin Actions ─────────────────────────────────────────────────────

@admin.action(description='▶︎  Launch selected campaigns')
def action_launch(modeladmin, request, queryset):
    for campaign in queryset:
        # ── Pre-flight checks ──────────────────────────────────────────────────
        if campaign.status == Campaign.STATUS_RUNNING:
            messages.warning(
                request,
                f'"{campaign.name}" skipped — already running.'
            )
            continue
        if not campaign.targets.exists():
            messages.warning(
                request,
                f'"{campaign.name}" skipped — no targets added. '
                f'Add targets via the Targets inline below.'
            )
            continue
        if not campaign.email_template:
            messages.warning(
                request,
                f'"{campaign.name}" skipped — no email template assigned. '
                f'Edit the campaign and assign a template.'
            )
            continue
        if not campaign.smtp_host or not campaign.smtp_user:
            messages.warning(
                request,
                f'"{campaign.name}" skipped — SMTP configuration incomplete. '
                f'Edit the campaign and fill in SMTP Host and Username.'
            )
            continue
        if not campaign.from_email:
            messages.warning(
                request,
                f'"{campaign.name}" skipped — From Email is not set. '
                f'Edit the campaign and fill in the From Email field.'
            )
            continue

        # ── Launch ─────────────────────────────────────────────────────────────
        try:
            from django_q.tasks import async_task
            from apps.campaigns.tasks import launch_campaign_async
            async_task(
                'apps.campaigns.tasks.launch_campaign_async',
                campaign.id,
                task_name=f'launch-campaign-{campaign.id}',
            )
            messages.success(
                request,
                f'"{campaign.name}" queued — '
                f'{campaign.targets.count()} email(s) will be sent by the worker. '
                f'Make sure the qcluster worker is running '
                f'(python manage.py qcluster).'
            )
        except Exception as q_err:
            # Django-Q2 unavailable or DB issue — fall back to synchronous send
            logger.warning(
                f'Django-Q unavailable ({q_err}), '
                f'launching "{campaign.name}" synchronously.'
            )
            try:
                from apps.campaigns.tasks import _launch_sync
                sent, failed, first_error = _launch_sync(campaign.id)
                if failed == 0:
                    messages.success(
                        request,
                        f'"{campaign.name}" launched synchronously — '
                        f'{sent} email(s) sent successfully.'
                    )
                elif sent == 0:
                    messages.error(
                        request,
                        f'"{campaign.name}" — all {failed} email(s) failed. '
                        f'Error: {first_error}. '
                        f'Check your SMTP credentials.'
                    )
                else:
                    messages.warning(
                        request,
                        f'"{campaign.name}" — {sent} sent, {failed} failed. '
                        f'First error: {first_error}.'
                    )
            except Exception as sync_err:
                messages.error(
                    request,
                    f'"{campaign.name}" failed to launch: {sync_err}'
                )


@admin.action(description='❚❚  Pause selected campaigns')
def action_pause(modeladmin, request, queryset):
    count = queryset.filter(status=Campaign.STATUS_RUNNING).update(
        status=Campaign.STATUS_PAUSED
    )
    messages.success(request, f'{count} campaign(s) paused.')


@admin.action(description='✓  Mark selected campaigns as completed')
def action_complete(modeladmin, request, queryset):
    count = 0
    for campaign in queryset:
        campaign.status       = Campaign.STATUS_COMPLETED
        campaign.completed_at = timezone.now()
        campaign.save(update_fields=['status', 'completed_at'])
        count += 1
    messages.success(request, f'{count} campaign(s) marked as completed.')


@admin.action(description='⟳  Reset selected campaigns to Draft')
def action_reset_draft(modeladmin, request, queryset):
    count = queryset.update(status=Campaign.STATUS_DRAFT)
    messages.success(request, f'{count} campaign(s) reset to Draft.')


# ── Campaign Admin ─────────────────────────────────────────────────────────────

@admin.register(Campaign)
class CampaignAdmin(admin.ModelAdmin):
    form = CampaignAdminForm
    list_display = (
        'name', 'status_badge', 'email_template',
        'total_targets_display', 'emails_sent_display',
        'links_clicked_display', 'click_rate_display',
        'created_by', 'created_at',
    )
    list_filter   = ('status', 'created_at')
    search_fields = ('name', 'description')
    readonly_fields = (
        'created_at', 'launched_at', 'completed_at',
        'total_targets_display', 'emails_sent_display',
        'links_clicked_display', 'lms_completed_display',
        'click_rate_display',
    )
    inlines = [CampaignTargetInline]
    actions = [action_launch, action_pause, action_complete, action_reset_draft]

    def save_model(self, request, obj, form, change):
        # Save the campaign first so it has a PK
        super().save_model(request, obj, form, change)

        # Now process the uploaded file if one was provided
        upload_file = request.FILES.get('csv_file')
        if not upload_file:
            return

        filename = upload_file.name.lower()
        rows = []

        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(upload_file, read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    headers = [
                        str(h).strip().lower() if h else ''
                        for h in next(rows_iter)
                    ]
                except StopIteration:
                    headers = []
                for row in rows_iter:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_dict[headers[i]] = str(val).strip() if val is not None else ''
                    rows.append(row_dict)
                wb.close()
            else:
                # CSV — read all bytes once to avoid the empty-read bug
                raw = upload_file.read()
                try:
                    text = raw.decode('utf-8-sig')
                except UnicodeDecodeError:
                    text = raw.decode('latin-1')
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    rows.append({
                        k.strip().lower(): (v or '')
                        for k, v in row.items()
                        if k
                    })
        except Exception as parse_err:
            self.message_user(
                request,
                f'Failed to parse uploaded file: {parse_err}',
                messages.ERROR,
            )
            return

        created = skipped = errors = 0
        for row in rows:
            email = row.get('email', '').strip().lower()
            if not email or '@' not in email:
                errors += 1
                continue
            _, was_created = CampaignTarget.objects.get_or_create(
                campaign=obj,
                email=email,
                defaults={
                    'full_name': row.get('full_name', '').strip(),
                    'department': row.get('department', '').strip(),
                    'position': row.get('position', '').strip(),
                    'business_unit': row.get('business_unit', '').strip(),
                    'manager': row.get('manager', '').strip(),
                    'manager_email': row.get('manager_email', '').strip(),
                },
            )
            if was_created:
                created += 1
            else:
                skipped += 1

        if created == 0 and skipped == 0 and errors == 0:
            self.message_user(
                request,
                'File was uploaded but contained no rows. '
                'Make sure it has an "email" header column.',
                messages.WARNING,
            )
        elif created == 0 and errors > 0:
            self.message_user(
                request,
                f'File uploaded but no valid emails found ({errors} invalid rows). '
                f'Make sure the column is named "email".',
                messages.WARNING,
            )
        else:
            msg = f'File upload: {created} target(s) added.'
            if skipped:
                msg += f' {skipped} duplicate(s) skipped.'
            if errors:
                msg += f' {errors} row(s) had missing/invalid emails.'
            self.message_user(request, msg, messages.SUCCESS)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                '<int:campaign_id>/upload-csv/',
                self.admin_site.admin_view(self.upload_csv_view),
                name='campaigns_campaign_upload_csv',
            ),
        ]
        return custom + urls

    def upload_csv_view(self, request, campaign_id):
        from django.http import HttpResponse
        from django.middleware.csrf import get_token

        try:
            campaign = Campaign.objects.get(pk=campaign_id)
        except Campaign.DoesNotExist:
            messages.error(request, 'Campaign not found.')
            return HttpResponseRedirect('/admin/campaigns/campaign/')

        if request.method == 'POST' and request.FILES.get('csv_file'):
            upload_file = request.FILES['csv_file']
            filename    = upload_file.name.lower()

            # Parse rows from CSV or XLSX
            rows = []
            try:
                if filename.endswith('.xlsx') or filename.endswith('.xls'):
                    wb = openpyxl.load_workbook(upload_file, read_only=True, data_only=True)
                    ws = wb.active
                    rows_iter = ws.iter_rows(values_only=True)
                    try:
                        headers = [
                            str(h).strip().lower() if h else ''
                            for h in next(rows_iter)
                        ]
                    except StopIteration:
                        headers = []
                    for row in rows_iter:
                        row_dict = {}
                        for i, val in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = str(val).strip() if val is not None else ''
                        rows.append(row_dict)
                    wb.close()
                else:
                    # CSV — read all bytes at once to avoid empty-read bug
                    raw = upload_file.read()
                    try:
                        text = raw.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        text = raw.decode('latin-1')
                    reader = csv.DictReader(io.StringIO(text))
                    for row in reader:
                        rows.append({
                            k.strip().lower(): (v or '') for k, v in row.items()
                        })
            except Exception as parse_err:
                messages.error(request, f'Failed to parse file: {parse_err}')
                rows = []

            created = skipped = errors = 0
            for row in rows:
                email = row.get('email', '').strip().lower()
                if not email or '@' not in email:
                    errors += 1
                    continue
                _, was_created = CampaignTarget.objects.get_or_create(
                    campaign=campaign,
                    email=email,
                    defaults={
                        'full_name': row.get('full_name', '').strip(),
                        'department': row.get('department', '').strip(),
                        'position': row.get('position', '').strip(),
                        'business_unit': row.get('business_unit', '').strip(),
                        'manager': row.get('manager', '').strip(),
                        'manager_email': row.get('manager_email', '').strip(),
                    },
                )
                if was_created:
                    created += 1
                else:
                    skipped += 1

            if created == 0 and skipped == 0:
                messages.warning(
                    request,
                    'No valid email addresses found. '
                    'Make sure the file has an "email" column header.'
                )
            else:
                msg = f'Uploaded: {created} target(s) added.'
                if skipped:
                    msg += f' {skipped} duplicate(s) skipped.'
                if errors:
                    msg += f' {errors} row(s) skipped (missing or invalid email).'
                messages.success(request, msg)

            return HttpResponseRedirect(
                f'/admin/campaigns/campaign/{campaign_id}/change/'
            )

        csrf_token = get_token(request)
        target_count = campaign.targets.count()

        html = (
            '<!DOCTYPE html><html><head>'
            '<title>Upload Targets CSV | PhishingOps Admin</title>'
            '<link rel="stylesheet" href="/static/admin/css/base.css">'
            '<style>'
            '#container{padding:20px;max-width:700px;}'
            'table{width:100%;border-collapse:collapse;margin-top:8px;}'
            'th,td{padding:6px 10px;border:1px solid #ddd;text-align:left;font-size:13px;}'
            'th{background:#f8f8f8;font-weight:bold;}'
            'code{background:#f0f0f0;padding:1px 5px;border-radius:3px;}'
            '.btn{background:#417690;color:white;padding:8px 20px;'
            'border:none;border-radius:4px;cursor:pointer;font-size:14px;}'
            '.btn:hover{background:#205067;}'
            '</style></head>'
            '<body class="default"><div id="container">'
            '<div id="header"><div id="branding">'
            '<h1><a href="/admin/">PhishingOps Administration</a></h1>'
            '</div></div>'
            '<div class="breadcrumbs">'
            '<a href="/admin/">Home</a> &rsaquo; '
            '<a href="/admin/campaigns/campaign/">Campaigns</a> &rsaquo; '
            f'<a href="/admin/campaigns/campaign/{campaign_id}/change/">{campaign.name}</a> &rsaquo; '
            'Upload Targets CSV'
            '</div>'
            '<div id="content" class="colM" style="padding-top:20px;">'
            f'<h1>Upload Targets CSV &mdash; {campaign.name}</h1>'
            f'<p>Current targets: <strong>{target_count}</strong></p>'
            '<fieldset style="border:1px solid #ccc;border-radius:4px;padding:16px 20px;margin-bottom:20px;max-width:500px;">'
            '<legend style="font-weight:bold;padding:0 8px;">CSV Format</legend>'
            '<p style="margin:0 0 10px;">First row must be a header. Column names are case-insensitive.</p>'
            '<table>'
            '<tr><th>Column</th><th>Required</th><th>Description</th></tr>'
            '<tr><td><code>email</code></td><td>&#10003; Yes</td><td>Employee email address</td></tr>'
            '<tr><td><code>full_name</code></td><td>No</td><td>Full name</td></tr>'
            '<tr><td><code>department</code></td><td>No</td><td>Department</td></tr>'
            '<tr><td><code>position</code></td><td>No</td><td>Job title</td></tr>'
            '<tr><td><code>business_unit</code></td><td>No</td><td>Business unit</td></tr>'
            '<tr><td><code>manager</code></td><td>No</td><td>Manager full name</td></tr>'
            '<tr><td><code>manager_email</code></td><td>No</td><td>Manager email address</td></tr>'
            '</table>'
            '<p style="margin-top:10px;font-size:12px;color:#666;">'
            'Example row:<br>'
            '<code>john@company.com,John Doe,Engineering,Developer</code>'
            '</p>'
            '</fieldset>'
            f'<form method="post" enctype="multipart/form-data">'
            f'<input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">'
            '<fieldset style="border:1px solid #ccc;border-radius:4px;padding:16px 20px;margin-bottom:20px;max-width:500px;">'
            '<legend style="font-weight:bold;padding:0 8px;">Select File</legend>'
            '<div style="margin-bottom:12px;">'
            '<label style="display:block;font-weight:bold;margin-bottom:6px;">CSV File:</label>'
            '<input type="file" name="csv_file" accept=".csv,.xlsx,.xls" required style="padding:4px;">'
            '</div>'
            '</fieldset>'
            '<div>'
            '<button type="submit" class="btn">Upload CSV</button>'
            '&nbsp;&nbsp;'
            f'<a href="/admin/campaigns/campaign/{campaign_id}/change/" style="color:#447e9b;">'
            '&larr; Back to Campaign</a>'
            '</div>'
            '</form>'
            '</div></div></body></html>'
        )

        return HttpResponse(html)

    fieldsets = (
        ('Campaign Details', {
            'fields': (
                'name', 'description', 'status',
                'email_template',
                'created_by', 'scheduled_at',
            )
        }),
        ('Timestamps', {
            'fields': ('created_at', 'launched_at', 'completed_at'),
            'classes': ('collapse',),
        }),
        ('Stats', {
            'fields': (
                'total_targets_display', 'emails_sent_display',
                'links_clicked_display', 'lms_completed_display',
                'click_rate_display',
            ),
            'classes': ('collapse',),
        }),
        ('SMTP Configuration', {
            'fields': (
                'smtp_host', 'smtp_port', 'smtp_user',
                'smtp_password', 'from_email',
                'smtp_use_tls', 'smtp_use_ssl',
            ),
            'description': 'SMTP password is stored encrypted.',
        }),
        ('Upload Targets (CSV)', {
            'fields': ('csv_file',),
            'description': (
                'Upload a CSV file to bulk-add targets. '
                "Required column: 'email' (case-insensitive). "
                "Optional: 'full_name', 'department', 'position', 'business_unit', 'manager', 'manager_email'. "
                'Duplicate emails are skipped.'
            ),
        }),
    )

    # ── Computed column displays ───────────────────────────────────────────────

    def status_badge(self, obj):
        colours = {
            'draft':     '#888',
            'running':   '#28a745',
            'paused':    '#fd7e14',
            'completed': '#6c757d',
        }
        colour = colours.get(obj.status, '#888')
        return format_html(
            '<span style="color:{}; font-weight:bold;">{}</span>',
            colour,
            obj.get_status_display(),
        )
    status_badge.short_description = 'Status'

    def total_targets_display(self, obj):
        return obj.total_targets
    total_targets_display.short_description = 'Targets'

    def emails_sent_display(self, obj):
        return obj.emails_sent
    emails_sent_display.short_description = 'Sent'

    def links_clicked_display(self, obj):
        return obj.links_clicked
    links_clicked_display.short_description = 'Clicked'

    def lms_completed_display(self, obj):
        return obj.lms_completed
    lms_completed_display.short_description = 'LMS Completed'

    def click_rate_display(self, obj):
        return f'{obj.click_rate}%'
    click_rate_display.short_description = 'Click Rate'


# ── Campaign Target Admin ──────────────────────────────────────────────────────

@admin.register(CampaignTarget)
class CampaignTargetAdmin(admin.ModelAdmin):
    list_display  = (
        'email', 'full_name', 'department', 'business_unit', 'campaign', 'manager', 'manager_email',
        'email_sent_at', 'clicked_display', 'lms_completed_at', 'quiz_score',
    )
    list_filter   = (
        'campaign', 'email_failed',
        'link_clicked_at', 'lms_completed_at',
    )
    search_fields = ('email', 'full_name', 'department', 'campaign__name')
    readonly_fields = (
        'token', 'phishing_link_display',
        'email_sent_at', 'email_failed', 'email_error',
        'link_clicked_at', 'click_ip', 'click_user_agent',
        'lms_started_at', 'lms_completed_at', 'quiz_score',
    )

    fieldsets = (
        ('Employee Info', {
            'fields': (
                'campaign', 'email', 'full_name', 'department',
                'position', 'business_unit', 'manager', 'manager_email',
            )
        }),
        ('Phishing Link', {
            'fields': ('token', 'phishing_link_display'),
        }),
        ('Email Tracking', {
            'fields': ('email_sent_at', 'email_failed', 'email_error'),
            'classes': ('collapse',),
        }),
        ('Click Tracking', {
            'fields': ('link_clicked_at', 'click_ip', 'click_user_agent'),
            'classes': ('collapse',),
        }),
        ('LMS Tracking', {
            'fields': ('lms_started_at', 'lms_completed_at', 'quiz_score'),
            'classes': ('collapse',),
        }),
    )

    def clicked_display(self, obj):
        if obj.link_clicked_at:
            return format_html(
                '<span style="color:red;">&#10003; {}</span>',
                localtime(obj.link_clicked_at).strftime('%Y-%m-%d %H:%M')
            )
        return '—'
    clicked_display.short_description = 'Clicked'

    def phishing_link_display(self, obj):
        link = obj.phishing_link
        return format_html('<a href="{}" target="_blank">{}</a>', link, link)
    phishing_link_display.short_description = 'Phishing Link'


# ── Progress / Attempt Admin ───────────────────────────────────────────────────

@admin.register(LessonProgress)
class LessonProgressAdmin(admin.ModelAdmin):
    list_display  = ('target', 'lesson', 'started_at', 'completed_at')
    list_filter   = ('lesson__course', 'completed_at')
    search_fields = ('target__email',)
    readonly_fields = ('started_at',)


@admin.register(QuizAttempt)
class QuizAttemptAdmin(admin.ModelAdmin):
    list_display  = ('target', 'quiz', 'score', 'passed', 'submitted_at')
    list_filter   = ('passed', 'quiz__course', 'submitted_at')
    search_fields = ('target__email',)
    readonly_fields = ('submitted_at', 'answers')
